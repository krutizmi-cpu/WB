import io
import json
import re
from pathlib import Path

import pandas as pd
import streamlit as st
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
CONFIG_PATH = ROOT / "config.json"
LOOKUP_PATH = DATA_DIR / "wb_kvv_lookup.csv"
TEMPLATE_PATH = ROOT / "templates" / "wb_template.xlsx"


@st.cache_data(show_spinner=False)
def load_config() -> dict:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


@st.cache_data(show_spinner=False)
def load_lookup() -> pd.DataFrame:
    return pd.read_csv(LOOKUP_PATH)


CFG = load_config()
LOOKUP = load_lookup()

REQUIRED_COLUMNS = [
    "Артикул",
    "Наименование",
    "Длина, см",
    "Ширина, см",
    "Высота, см",
]

OPTIONAL_COLUMNS = [
    "Вес, кг",
    "Себестоимость, ₽",
    "Цена продавца до СПП, ₽",
]

PREFERRED_SUBJECT_PATTERNS = {
    "велосипед": [r"\bвелосипеды\b", r"\bвелосипеды двухколесные\b", r"\bвелосипеды трехколесные\b"],
    "дрель": [r"\bдрели\b"],
    "шуруповерт": [r"\bшуруповерты\b"],
    "игрушка": [r"\bмягкие игрушки\b", r"\bигрушки\b"],
    "платье": [r"\bплатья\b"],
    "стол": [r"\bстолы письменные\b", r"\bстолы\b"],
    "шампунь": [r"\bшампуни\b"],
}

KEYWORD_OVERRIDES = [
    ("велосипед", ["велосипед"]),
    ("дрель", ["дрел"]),
    ("шуруповерт", ["шуруповерт"]),
    ("перфоратор", ["перфоратор"]),
    ("сверло", ["сверл"]),
    ("игрушка", ["игрушк"]),
    ("платье", ["плать"]),
    ("куртка", ["куртк"]),
    ("футболка", ["футболк"]),
    ("джинсы", ["джинс"]),
    ("брюки", ["брюк"]),
    ("кроссовки", ["кроссовк"]),
    ("ботинки", ["ботин"]),
    ("ноутбук", ["ноутбук"]),
    ("смартфон", ["смартфон"]),
    ("наушники", ["наушник"]),
    ("чехол", ["чехл"]),
    ("кастрюля", ["кастрюл"]),
    ("сковорода", ["сковород"]),
    ("кружка", ["кружк"]),
    ("стол", ["стол"]),
    ("стул", ["стул"]),
    ("диван", ["диван"]),
    ("шкаф", ["шкаф"]),
    ("матрас", ["матрас"]),
    ("подгузники", ["подгуз"]),
    ("корм", ["корм"]),
    ("наполнитель", ["наполнител"]),
    ("шампунь", ["шампун"]),
    ("крем", ["крем"]),
]


def normalize_text(value: str) -> str:
    value = str(value or "").lower().replace("ё", "е")
    value = re.sub(r"[^a-zа-я0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "sku": "Артикул",
        "артикул продавца": "Артикул",
        "наименование товара": "Наименование",
        "товар": "Наименование",
        "длина": "Длина, см",
        "ширина": "Ширина, см",
        "высота": "Высота, см",
        "вес": "Вес, кг",
        "себес": "Себестоимость, ₽",
        "себестоимость": "Себестоимость, ₽",
        "цена": "Цена продавца до СПП, ₽",
        "цена продавца": "Цена продавца до СПП, ₽",
    }
    renamed = {}
    for col in df.columns:
        key = normalize_text(col)
        if key in aliases:
            renamed[col] = aliases[key]
    df = df.rename(columns=renamed)

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(
            "В файле не хватает обязательных колонок: " + ", ".join(missing)
        )

    for col in OPTIONAL_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df


def numeric_series(df: pd.DataFrame, col: str, default: float) -> pd.Series:
    return pd.to_numeric(df[col], errors="coerce").fillna(default)


@st.cache_data(show_spinner=False)
def build_index() -> list:
    items = []
    for _, row in LOOKUP.iterrows():
        items.append(
            {
                "lookup_text": str(row["lookup_text"]),
                "lookup_norm": str(row.get("lookup_norm", "")),
                "category_name": str(row.get("category_name", "")),
                "category_norm": str(row.get("category_norm", "")),
                "subject_name": str(row.get("subject_name", "")),
                "subject_norm": str(row.get("subject_norm", "")),
                "fbw_percent": float(row["fbw_percent"]),
                "fbs_percent": float(row["fbs_percent"]),
            }
        )
    return items


INDEX = build_index()


def extract_commission_percent(record: dict) -> float:
    model = str(CFG.get("model", "FBW")).upper()
    return record["fbw_percent"] if model == "FBW" else record["fbs_percent"]


def choose_override_record(norm_name: str):
    candidates = []

    for label, stems in KEYWORD_OVERRIDES:
        if not any(stem in norm_name for stem in stems):
            continue

        preferred_patterns = PREFERRED_SUBJECT_PATTERNS.get(label, [])
        if preferred_patterns:
            preferred_hits = []
            for item in INDEX:
                subj = item["subject_norm"]
                if any(re.search(pattern, subj) for pattern in preferred_patterns):
                    score = max(
                        fuzz.token_set_ratio(norm_name, subj),
                        fuzz.partial_ratio(norm_name, subj),
                        fuzz.token_set_ratio(norm_name, item["lookup_norm"]),
                    )
                    preferred_hits.append((score + 20, item))
            if preferred_hits:
                preferred_hits.sort(key=lambda x: x[0], reverse=True)
                return preferred_hits[0][1], preferred_hits[0][0]

        for item in INDEX:
            haystack = f"{item['subject_norm']} {item['lookup_norm']} {item['category_norm']}"
            if any(stem in haystack for stem in stems):
                score = max(
                    fuzz.token_set_ratio(norm_name, item["subject_norm"]),
                    fuzz.partial_ratio(norm_name, item["subject_norm"]),
                    fuzz.token_set_ratio(norm_name, item["lookup_norm"]),
                )
                if label in item["subject_norm"]:
                    score += 15
                candidates.append((score, item))

    if not candidates:
        return None, None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1], candidates[0][0]


def classify_name(name: str) -> dict:
    norm_name = normalize_text(name)
    if not norm_name:
        return {
            "Категория WB": "Не определена",
            "Предмет WB": "Не определен",
            "Источник комиссии": "fallback",
            "Комиссия WB, %": float(CFG["fallback_commission_percent"]),
            "Скоринг категории": 0,
        }

    override_record, override_score = choose_override_record(norm_name)
    if override_record is not None and override_score >= 78:
        return {
            "Категория WB": override_record["category_name"] or override_record["lookup_text"],
            "Предмет WB": override_record["subject_name"] or override_record["lookup_text"],
            "Источник комиссии": "wb_kvv_pdf_2025-12-22_keyword",
            "Комиссия WB, %": round(extract_commission_percent(override_record), 2),
            "Скоринг категории": round(override_score, 1),
        }

    best_score = -1
    best_record = None

    for item in INDEX:
        subject_score = max(
            fuzz.token_set_ratio(norm_name, item["subject_norm"]),
            fuzz.partial_ratio(norm_name, item["subject_norm"]),
        )
        lookup_score = fuzz.token_set_ratio(norm_name, item["lookup_norm"])
        score = max(subject_score, lookup_score * 0.92)

        if item["subject_norm"] and (item["subject_norm"] in norm_name or norm_name in item["subject_norm"]):
            score = max(score, 90)
        if item["lookup_norm"] and (item["lookup_norm"] in norm_name or norm_name in item["lookup_norm"]):
            score = max(score, 88)

        if score > best_score:
            best_score = score
            best_record = item

    if best_record is None or best_score < float(CFG["match_threshold"]):
        return {
            "Категория WB": "Не определена",
            "Предмет WB": "Не определен",
            "Источник комиссии": "fallback",
            "Комиссия WB, %": float(CFG["fallback_commission_percent"]),
            "Скоринг категории": round(best_score if best_score > 0 else 0, 1),
        }

    return {
        "Категория WB": best_record["category_name"] or best_record["lookup_text"],
        "Предмет WB": best_record["subject_name"] or best_record["lookup_text"],
        "Источник комиссии": "wb_kvv_pdf_2025-12-22",
        "Комиссия WB, %": round(extract_commission_percent(best_record), 2),
        "Скоринг категории": round(best_score, 1),
    }


def calc_tax(price: pd.Series) -> pd.Series:
    tax_mode = str(CFG.get("tax_mode", "USN_6")).upper()
    if tax_mode == "OSNO":
        return price * float(CFG["tax_percent_osno"]) / 100.0
    return price * float(CFG["tax_percent_usn"]) / 100.0


def calculate(df: pd.DataFrame) -> pd.DataFrame:
    df = ensure_columns(df.copy())

    df["Артикул"] = df["Артикул"].astype(str).str.strip()
    df["Наименование"] = df["Наименование"].astype(str).str.strip()

    df["Длина, см"] = numeric_series(df, "Длина, см", 0.0)
    df["Ширина, см"] = numeric_series(df, "Ширина, см", 0.0)
    df["Высота, см"] = numeric_series(df, "Высота, см", 0.0)
    df["Вес, кг"] = numeric_series(df, "Вес, кг", float(CFG["default_weight_kg"]))
    df["Себестоимость, ₽"] = numeric_series(df, "Себестоимость, ₽", float(CFG["default_cogs_rub"]))
    df["Цена продавца до СПП, ₽"] = numeric_series(df, "Цена продавца до СПП, ₽", float(CFG["default_seller_price_rub"]))

    classified = df["Наименование"].apply(classify_name).apply(pd.Series)
    df = pd.concat([df, classified], axis=1)

    df["Объем, л"] = (df["Длина, см"] * df["Ширина, см"] * df["Высота, см"]) / 1000.0
    df["Логистика базовая, ₽"] = (
        float(CFG["logistics_base_rub_per_item_up_to_1l"])
        + df["Объем, л"].sub(1).clip(lower=0) * float(CFG["logistics_additional_rub_per_liter_over_1l"])
    ) * float(CFG["warehouse_logistics_coefficient"])

    df["Хранение, ₽"] = (
        df["Объем, л"]
        * float(CFG["storage_rub_per_liter_per_day"])
        * float(CFG["storage_days"])
        * float(CFG["warehouse_storage_coefficient"])
    )
    df["Приемка, ₽"] = float(CFG["acceptance_rub_per_item"])

    df["Комиссия WB, ₽"] = df["Цена продавца до СПП, ₽"] * df["Комиссия WB, %"] / 100.0

    spp_rate = float(CFG["spp_percent"]) / 100.0
    df["СПП, %"] = float(CFG["spp_percent"])
    df["Цена для покупателя после СПП, ₽"] = df["Цена продавца до СПП, ₽"] * (1.0 - spp_rate)
    df["Сумма СПП, ₽"] = df["Цена продавца до СПП, ₽"] - df["Цена для покупателя после СПП, ₽"]

    df["Реклама, %"] = float(CFG["ad_percent"])
    df["Реклама, ₽"] = df["Цена продавца до СПП, ₽"] * float(CFG["ad_percent"]) / 100.0

    df["Эквайринг, %"] = float(CFG["acquiring_percent"])
    df["Эквайринг, ₽"] = df["Цена продавца до СПП, ₽"] * float(CFG["acquiring_percent"]) / 100.0

    buyout_rate = max(float(CFG["buyout_percent"]) / 100.0, 0.01)
    df["Выкуп, %"] = float(CFG["buyout_percent"])
    df["Ожидаемая стоимость возвратов/невыкупов, ₽"] = (
        ((1.0 - buyout_rate) / buyout_rate) * float(CFG["return_logistics_rub"])
    )

    defect_rate = float(CFG["defect_percent"]) / 100.0
    defect_loss = float(CFG["defect_loss_percent_of_cogs"]) / 100.0
    df["Брак, %"] = float(CFG["defect_percent"])
    df["Ожидаемая стоимость брака, ₽"] = df["Себестоимость, ₽"] * defect_rate * defect_loss

    df["Налог, ₽"] = calc_tax(df["Цена продавца до СПП, ₽"])

    cost_cols = [
        "Себестоимость, ₽",
        "Комиссия WB, ₽",
        "Логистика базовая, ₽",
        "Хранение, ₽",
        "Приемка, ₽",
        "Реклама, ₽",
        "Эквайринг, ₽",
        "Ожидаемая стоимость возвратов/невыкупов, ₽",
        "Ожидаемая стоимость брака, ₽",
        "Налог, ₽",
    ]
    df["Полная себестоимость, ₽"] = df[cost_cols].sum(axis=1)

    df["Прибыль, ₽"] = df["Цена продавца до СПП, ₽"] - df["Полная себестоимость, ₽"]
    df["Маржа % от цены"] = df["Прибыль, ₽"] / df["Цена продавца до СПП, ₽"].replace(0, pd.NA)
    df["Наценка %"] = (
        df["Цена продавца до СПП, ₽"] / df["Полная себестоимость, ₽"].replace(0, pd.NA) - 1.0
    )

    output_order = [
        "Артикул",
        "Наименование",
        "Категория WB",
        "Предмет WB",
        "Скоринг категории",
        "Источник комиссии",
        "Длина, см",
        "Ширина, см",
        "Высота, см",
        "Вес, кг",
        "Объем, л",
        "Цена продавца до СПП, ₽",
        "СПП, %",
        "Цена для покупателя после СПП, ₽",
        "Сумма СПП, ₽",
        "Комиссия WB, %",
        "Комиссия WB, ₽",
        "Логистика базовая, ₽",
        "Хранение, ₽",
        "Приемка, ₽",
        "Реклама, %",
        "Реклама, ₽",
        "Эквайринг, %",
        "Эквайринг, ₽",
        "Выкуп, %",
        "Ожидаемая стоимость возвратов/невыкупов, ₽",
        "Брак, %",
        "Ожидаемая стоимость брака, ₽",
        "Налог, ₽",
        "Себестоимость, ₽",
        "Полная себестоимость, ₽",
        "Прибыль, ₽",
        "Маржа % от цены",
        "Наценка %",
    ]
    return df[output_order]


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()
    ws = wb.active
    ws.title = "WB unit economics"

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1F2")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    for row in df.itertuples(index=False):
        ws.append(list(row))

    percent_cols = {"СПП, %", "Комиссия WB, %", "Реклама, %", "Эквайринг, %", "Выкуп, %", "Брак, %", "Маржа % от цены", "Наценка %"}
    currency_cols = {c for c in df.columns if "₽" in c}
    number_cols = {"Длина, см", "Ширина, см", "Высота, см", "Вес, кг", "Объем, л", "Скоринг категории"}

    for idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        width = max(12, min(28, int(max(len(str(col_name)), 12) * 1.1)))
        if col_name in {"Наименование", "Категория WB", "Предмет WB", "Источник комиссии"}:
            width = 32
        ws.column_dimensions[letter].width = width

        for cell in ws[letter][1:]:
            if col_name in currency_cols:
                cell.number_format = '#,##0.00;[Red](#,##0.00);-'
            elif col_name in percent_cols:
                if col_name in {"Маржа % от цены", "Наценка %"}:
                    cell.number_format = '0.0%'
                else:
                    cell.number_format = '0.0'
            elif col_name in number_cols:
                cell.number_format = '0.00'
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    last_row = ws.max_row
    if last_row >= 2:
        profit_col = df.columns.get_loc("Прибыль, ₽") + 1
        ws.conditional_formatting.add(
            f"{get_column_letter(profit_col)}2:{get_column_letter(profit_col)}{last_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FDE9E7")),
        )

    notes = wb.create_sheet("README")
    notes["A1"] = "Что важно"
    notes["A1"].font = Font(bold=True, size=14)
    notes["A3"] = "1. Обязательные колонки: Артикул, Наименование, Длина, Ширина, Высота."
    notes["A4"] = "2. Вес, себестоимость и цена продавца — необязательные. Если их нет, берутся значения из config.json."
    notes["A5"] = "3. Комиссия определяется автоматически по сопоставлению названия товара со справочником WB КВВ."
    notes["A6"] = "4. Формулы выкупа, брака, СПП и налогов заданы как управляемые допущения в config.json."
    notes["A7"] = "5. Источники: seller.wildberries.ru раздел «Тарифы» и PDF таблица КВВ."
    notes.column_dimensions["A"].width = 120

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


st.set_page_config(page_title="WB Unit Economics", page_icon="🫐", layout="wide")
st.title("WB Unit Economics")
st.caption("Массовая загрузка товаров Excel → автоопределение категории → расчёт юнит-экономики → выгрузка Excel")

with open(TEMPLATE_PATH, "rb") as f:
    st.download_button(
        "Скачать шаблон Excel",
        data=f.read(),
        file_name="wb_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

uploaded_file = st.file_uploader(
    "Загрузите Excel с товарами",
    type=["xlsx", "xlsm", "xls"],
    accept_multiple_files=False,
)

with st.expander("Какие допущения зашиты в расчёт", expanded=False):
    st.write({
        "Модель": CFG["model"],
        "СПП, %": CFG["spp_percent"],
        "Реклама, %": CFG["ad_percent"],
        "Выкуп, %": CFG["buyout_percent"],
        "Брак, %": CFG["defect_percent"],
        "Потери от брака, % от себестоимости": CFG["defect_loss_percent_of_cogs"],
        "Эквайринг, %": CFG["acquiring_percent"],
        "Налоговый режим": CFG["tax_mode"],
        "Дней хранения": CFG["storage_days"],
        "Логистика до 1 литра, ₽": CFG["logistics_base_rub_per_item_up_to_1l"],
        "За каждый литр свыше 1, ₽": CFG["logistics_additional_rub_per_liter_over_1l"],
        "Хранение ₽/литр/день": CFG["storage_rub_per_liter_per_day"],
        "Возврат / невыкуп, ₽": CFG["return_logistics_rub"],
        "Fallback комиссия, %": CFG["fallback_commission_percent"],
    })

if uploaded_file is not None:
    try:
        raw_df = pd.read_excel(uploaded_file)
        result_df = calculate(raw_df)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("SKU", int(len(result_df)))
        c2.metric("Категория определена", f"{(result_df['Категория WB'] != 'Не определена').mean() * 100:.1f}%")
        c3.metric("Средняя комиссия, %", f"{result_df['Комиссия WB, %'].mean():.2f}")
        c4.metric("Средняя прибыль, ₽", f"{result_df['Прибыль, ₽'].mean():,.2f}".replace(",", " "))

        st.dataframe(result_df, use_container_width=True, height=560)

        excel_bytes = to_excel_bytes(result_df)
        st.download_button(
            "Скачать результат Excel",
            data=excel_bytes,
            file_name="wb_unit_economics_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as exc:
        st.error(f"Ошибка обработки файла: {exc}")
